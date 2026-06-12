"""
Módulo de Parsing e Normalização de Planilhas Excel.

Funcionalidades:
- Identificação automática de tipo de projeto (ETA, Captação, ETE)
- Leitura e normalização de dados de planilhas
- Extração de datas do nome do arquivo (formato YYYYMMDD)
- Forward fill de colunas agrupadas (Município Filtro, Município, Unidade, etc.)
- Tratamento de edge cases (typos, #VALUE!, células vazias, merged cells)
- Cache com @st.cache_data para performance

Tipos de projeto suportados:
- ETAs: Aba "Resumo e Plano Ação", 344 linhas, 11 municípios
- Captações: Aba "Plano Ação", 984 linhas, 23 grupos
- ETEs: Aba "Plano Ação", 2146 linhas, 40 localidades, 27 unidades
"""

import pandas as pd
import numpy as np
import openpyxl
import streamlit as st
from datetime import datetime
import re
from typing import Tuple, Optional, List
import logging

logger = logging.getLogger(__name__)


# Colunas base normalizadas (todas as planilhas)
COLUNAS_BASE = [
    "municipio_filtro",
    "municipio",
    "unidade",
    "atividade",
    "responsavel",
    "plano_de_acao",
    "evolucao",
    "status_atual",
    "data_entrega_final",
    "data_estimada",
    "data_conclusao",
    "tipo_projeto",
]

# Colunas exclusivas por tipo
COLUNAS_ETA = ["data_inicio_pe", "data_termino_pe", "status_da_eta"]
COLUNAS_ETE = ["observacoes"]
COLUNAS_CAPTACAO = []


def identificar_tipo(entrada: str | bytes) -> Optional[str]:
    """
    Identifica o tipo de projeto baseado no nome das abas da planilha.

    Aceita caminho de arquivo (str) ou bytes do arquivo.

    Lógica:
    - Se aba "CRONOGRAMA-FÍSICO" → None (Gantt incompatível)
    - Se aba "Resumo e Plano Ação" → "ETA"
    - Se aba "Plano Ação" + coluna "Unidade" → "ETE"
    - Se aba "Plano Ação" sem coluna "Unidade" → "Captação"

    Args:
        entrada: Caminho do arquivo .xlsx ou bytes do arquivo

    Returns:
        "ETA" | "Captação" | "ETE" | None (se incompatível ou erro)
    """
    try:
        # Se for bytes, converter para BytesIO
        if isinstance(entrada, bytes):
            import io
            entrada = io.BytesIO(entrada)

        wb = openpyxl.load_workbook(entrada, data_only=False)
        sheets_lower = [s.lower() for s in wb.sheetnames]

        # Verificar arquivo Gantt (incompatível)
        if any("cronograma" in s for s in sheets_lower):
            logger.warning(f"Arquivo Gantt detectado")
            return None

        # Verificar ETAs
        if any("resumo" in s and "plano" in s for s in sheets_lower):
            logger.info(f"Tipo ETA identificado")
            return "ETA"

        # Verificar Captações vs ETEs
        if any("plano ação" in s or "plano acao" in s for s in sheets_lower):
            try:
                # Ler header para verificar coluna "Unidade"
                for sheet in wb.sheetnames:
                    if "plano" in sheet.lower() and ("ação" in sheet.lower() or "acao" in sheet.lower()):
                        df = pd.read_excel(entrada, sheet_name=sheet, header=1, nrows=1)

                        if "Unidade" in df.columns:
                            logger.info(f"Tipo ETE identificado")
                            return "ETE"
                        else:
                            logger.info(f"Tipo Captação identificado")
                            return "Captação"
            except Exception as e:
                logger.error(f"Erro ao verificar Unidade: {e}")
                return None

        logger.warning(f"Tipo não identificado")
        return None

    except Exception as e:
        logger.error(f"Erro ao identificar tipo: {e}")
        return None


def _identificar_tipo_caminho(caminho_arquivo: str) -> Optional[str]:
    """
    Identifica o tipo de projeto baseado no nome das abas da planilha.

    Lógica:
    - Se aba "CRONOGRAMA-FÍSICO" → None (Gantt incompatível)
    - Se aba "Resumo e Plano Ação" → "ETA"
    - Se aba "Plano Ação" + coluna "Unidade" → "ETE"
    - Se aba "Plano Ação" sem coluna "Unidade" → "Captação"

    Args:
        caminho_arquivo: Caminho do arquivo .xlsx

    Returns:
        "ETA" | "Captação" | "ETE" | None (se incompatível ou erro)
    """
    try:
        wb = openpyxl.load_workbook(caminho_arquivo, data_only=False)
        sheets_lower = [s.lower() for s in wb.sheetnames]

        # Verificar arquivo Gantt (incompatível)
        if any("cronograma" in s for s in sheets_lower):
            logger.warning(f"Arquivo Gantt detectado: {caminho_arquivo}")
            return None

        # Verificar ETAs
        if any("resumo" in s and "plano" in s for s in sheets_lower):
            logger.info(f"Tipo ETA identificado: {caminho_arquivo}")
            return "ETA"

        # Verificar Captações vs ETEs
        if any("plano ação" in s or "plano acao" in s for s in sheets_lower):
            try:
                # Ler header para verificar coluna "Unidade"
                for sheet in wb.sheetnames:
                    if "plano" in sheet.lower() and "ação" in sheet.lower() or "acao" in sheet.lower():
                        df = pd.read_excel(caminho_arquivo, sheet_name=sheet, header=1, nrows=1)

                        if "Unidade" in df.columns:
                            logger.info(f"Tipo ETE identificado: {caminho_arquivo}")
                            return "ETE"
                        else:
                            logger.info(f"Tipo Captação identificado: {caminho_arquivo}")
                            return "Captação"
            except Exception as e:
                logger.error(f"Erro ao verificar Unidade: {e}")
                return None

        logger.warning(f"Tipo não identificado: {caminho_arquivo}")
        return None

    except Exception as e:
        logger.error(f"Erro ao identificar tipo: {e}")
        return None


def extrair_data_arquivo(nome_arquivo: str) -> Optional[datetime]:
    """
    Extrai data do nome do arquivo no formato YYYYMMDD.

    Exemplos:
    - "20260602 - Plano-de-Ação_Captações.xlsx" → datetime(2026, 6, 2)
    - "20260424 - Plano-de-Ação_ETAs.xlsx" → datetime(2026, 4, 24)

    Args:
        nome_arquivo: Nome do arquivo ou caminho

    Returns:
        datetime object ou None se não encontrar padrão YYYYMMDD
    """
    # Extrair primeiros 8 dígitos (YYYYMMDD)
    match = re.search(r"(\d{8})", nome_arquivo)
    if match:
        date_str = match.group(1)
        try:
            data = datetime.strptime(date_str, "%Y%m%d")
            logger.info(f"Data extraída: {date_str} → {data}")
            return data
        except ValueError:
            logger.warning(f"Data inválida: {date_str}")
            return None

    logger.warning(f"Padrão YYYYMMDD não encontrado em: {nome_arquivo}")
    return None


def normalizar_evolucao(valor) -> float:
    """
    Converte valor de Evolução para float [0, 1] ou NaN.

    Mapeamento:
    - 1.0 → 100% (concluído)
    - 0.8, 0.9, 0.5 → progresso parcial
    - 0.0 → 0% (pendente)
    - "-", "", None, NaN → NaN (não aplicável)

    Args:
        valor: Valor bruto da célula

    Returns:
        float entre 0 e 1, ou NaN
    """
    if valor == "-" or valor == "" or valor is None or pd.isna(valor):
        return float("nan")

    try:
        val_float = float(valor)
        # Validar intervalo [0, 1]
        if 0 <= val_float <= 1:
            return val_float
        else:
            logger.warning(f"Evolução fora do intervalo [0, 1]: {val_float}")
            return float("nan")
    except (ValueError, TypeError):
        logger.warning(f"Erro ao converter Evolução: {valor}")
        return float("nan")


def _normalizar_eta(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza planilha de ETA."""
    # Mapeamento de colunas
    mapa_colunas = {
        "Município Filtro": "municipio_filtro",
        "Município": "municipio",
        "Data Prevista de Início do PE": "data_inicio_pe",
        "Data de Prevista de Término do PE": "data_termino_pe",
        "Status Da ETA": "status_da_eta",
        "Data de Entrega Final": "data_entrega_final",
        "Atividade": "atividade",
        "Responsável": "responsavel",
        "Plano de Ação": "plano_de_acao",
        "Evolução": "evolucao",
        "Status Atual": "status_atual",
        "Data Estimada": "data_estimada",
        "Data de Conclusão": "data_conclusao",
    }

    # Renomear colunas existentes
    df = df.rename(columns=mapa_colunas)

    # Forward fill para Município, datas de PE e status da ETA
    df["municipio"] = df["municipio"].ffill()
    df["data_inicio_pe"] = df["data_inicio_pe"].ffill()
    df["data_termino_pe"] = df["data_termino_pe"].ffill()
    df["status_da_eta"] = df["status_da_eta"].ffill()
    df["data_entrega_final"] = df["data_entrega_final"].ffill()

    # Normalizar Evolução
    df["evolucao"] = df["evolucao"].apply(normalizar_evolucao)

    # Adicionar tipo
    df["tipo_projeto"] = "ETA"

    # Adicionar coluna Unidade (não aplicável para ETAs)
    if "unidade" not in df.columns:
        df["unidade"] = np.nan

    return df


def _normalizar_captacao(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza planilha de Captação."""
    # Mapeamento de colunas (com tratamento de typo "Responável")
    mapa_colunas = {
        "Município Filtro": "municipio_filtro",
        "Município": "municipio",
        "Data de Entrega Final": "data_entrega_final",
        "Atividade": "atividade",
        "Responável": "responsavel",  # Typo na planilha original
        "Responsável": "responsavel",  # Fallback se não tiver typo
        "Plano de Ação": "plano_de_acao",
        "Evolução": "evolucao",
        "Status Atual": "status_atual",
        "Data Estimada": "data_estimada",
        "Data de Conclusão": "data_conclusao",
    }

    # Renomear colunas existentes
    for col_original, col_novo in mapa_colunas.items():
        if col_original in df.columns:
            df = df.rename(columns={col_original: col_novo})

    # Forward fill para Município e Data de Entrega Final
    df["municipio"] = df["municipio"].ffill()
    df["data_entrega_final"] = df["data_entrega_final"].ffill()

    # Tratar erros em datas (ex: #VALUE!)
    df["data_estimada"] = pd.to_datetime(df["data_estimada"], errors="coerce")
    df["data_conclusao"] = pd.to_datetime(df["data_conclusao"], errors="coerce")

    # Normalizar Evolução
    df["evolucao"] = df["evolucao"].apply(normalizar_evolucao)

    # Adicionar tipo
    df["tipo_projeto"] = "Captação"

    # Adicionar coluna Unidade (não aplicável)
    if "unidade" not in df.columns:
        df["unidade"] = np.nan

    return df


def _normalizar_ete(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza planilha de ETE."""
    # Mapeamento de colunas (com tratamento de typo "Responável")
    mapa_colunas = {
        "Município Filtro": "municipio_filtro",
        "Município": "municipio",
        "Unidade": "unidade",
        "Data de Entrega Final": "data_entrega_final",
        "Atividade": "atividade",
        "Responável": "responsavel",  # Typo na planilha original
        "Responsável": "responsavel",  # Fallback
        "Plano de Ação": "plano_de_acao",
        "Evolução": "evolucao",
        "Status Atual": "status_atual",
        "Data Estimada": "data_estimada",
        "Data de Conclusão": "data_conclusao",
        "OBSERVAÇÕES": "observacoes",
        "Observações": "observacoes",
    }

    # Renomear colunas existentes
    for col_original, col_novo in mapa_colunas.items():
        if col_original in df.columns:
            df = df.rename(columns={col_original: col_novo})

    # Forward fill obrigatório para colunas mescladas
    # Município Filtro, Município, Unidade, Data Entrega Final, OBSERVAÇÕES
    df["municipio_filtro"] = df["municipio_filtro"].ffill()
    df["municipio"] = df["municipio"].ffill()
    df["unidade"] = df["unidade"].ffill()
    df["data_entrega_final"] = df["data_entrega_final"].ffill()
    df["observacoes"] = df["observacoes"].ffill()

    # Normalizar Evolução
    df["evolucao"] = df["evolucao"].apply(normalizar_evolucao)

    # Adicionar tipo
    df["tipo_projeto"] = "ETE"

    return df


@st.cache_data
def normalizar_planilha(entrada: "str | bytes", tipo: str) -> pd.DataFrame:
    """
    Normaliza planilha bruta para estrutura padrão.

    Args:
        entrada: Caminho do arquivo .xlsx (str) OU bytes do arquivo
        tipo: "ETA" | "Captação" | "ETE"

    Returns:
        DataFrame normalizado com colunas padrão
    """
    try:
        # Converter bytes para BytesIO para pd.read_excel
        if isinstance(entrada, (bytes, bytearray)):
            import io as _io
            fonte = _io.BytesIO(entrada)
        else:
            fonte = entrada  # str ou path

        # Determinar linha do header
        if tipo == "ETA":
            header_row = 2  # Linha 3 (0-indexed)
            sheet_name = "Resumo e Plano Ação"
        elif tipo == "Captação":
            header_row = 1  # Linha 2
            sheet_name = "Plano Ação"
        elif tipo == "ETE":
            header_row = 1  # Linha 2
            sheet_name = "Plano Ação"
        else:
            raise ValueError(f"Tipo desconhecido: {tipo}")

        # Ler planilha
        df = pd.read_excel(fonte, sheet_name=sheet_name, header=header_row)

        logger.info(f"Planilha lida: {tipo}, {len(df)} linhas, {len(df.columns)} colunas")

        # Normalizar conforme tipo
        if tipo == "ETA":
            df = _normalizar_eta(df)
        elif tipo == "Captação":
            df = _normalizar_captacao(df)
        elif tipo == "ETE":
            df = _normalizar_ete(df)

        # Garantir colunas base
        for col in COLUNAS_BASE:
            if col not in df.columns:
                df[col] = np.nan

        # Reordenar colunas
        colunas_finais = COLUNAS_BASE.copy()
        if tipo == "ETA":
            colunas_finais.extend(COLUNAS_ETA)
        elif tipo == "ETE":
            colunas_finais.extend(COLUNAS_ETE)

        # Selecionar apenas colunas necessárias
        df = df[[col for col in colunas_finais if col in df.columns]]

        logger.info(f"Planilha normalizada: {tipo}, {len(df)} linhas, {len(df.columns)} colunas")
        return df

    except Exception as e:
        logger.error(f"Erro ao normalizar planilha: {e}")
        raise


def processar_arquivo(arquivo_bytes: bytes, nome_arquivo: str) -> Tuple[Optional[pd.DataFrame], Optional[str], Optional[datetime]]:
    """
    Processa arquivo Excel enviado pelo usuário.

    Passos:
    1. Identificar tipo
    2. Extrair data
    3. Normalizar
    4. Retornar (df, tipo, data)

    Args:
        arquivo_bytes: Conteúdo do arquivo em bytes
        nome_arquivo: Nome do arquivo

    Returns:
        Tupla (DataFrame normalizado, tipo, data_extraída) ou (None, erro_msg, None)
    """
    import tempfile
    import os

    try:
        # Salvar temporariamente
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(arquivo_bytes)
            tmp_path = tmp.name

        # Identificar tipo
        tipo = identificar_tipo(tmp_path)
        if tipo is None:
            return None, "Arquivo incompatível ou tipo não identificado", None

        # Extrair data
        data = extrair_data_arquivo(nome_arquivo)

        # Normalizar
        df = normalizar_planilha(tmp_path, tipo)

        # Limpar temp
        os.remove(tmp_path)

        return df, tipo, data

    except Exception as e:
        logger.error(f"Erro ao processar arquivo: {e}")
        return None, str(e), None


def validar_dataframe(df: pd.DataFrame, tipo: str) -> Tuple[bool, str]:
    """
    Valida DataFrame normalizado.

    Args:
        df: DataFrame normalizado
        tipo: Tipo do projeto

    Returns:
        Tupla (é_válido, mensagem)
    """
    # Verificar colunas essenciais
    essenciais = ["municipio", "atividade", "status_atual"]
    for col in essenciais:
        if col not in df.columns or df[col].isna().all():
            return False, f"Coluna essencial vazia ou faltante: {col}"

    # Verificar linhas
    if len(df) == 0:
        return False, "DataFrame vazio"

    # Verificar Unidade para ETEs
    if tipo == "ETE" and (df["unidade"].isna().all()):
        return False, "Coluna 'Unidade' vazia para ETE"

    return True, "DataFrame válido"
